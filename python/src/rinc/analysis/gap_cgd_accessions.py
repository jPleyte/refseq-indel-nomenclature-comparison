'''
See which CGD transcripts are known to have refseq mismatch. 

1) Compare file1 with file2

2) Because a lot of the transcripts (1918 of them) exported from CGD are accession-only (no version) I also want to do this comparison with the ".x" version trimmed
Created on Jan 23, 2026

This tool is not currently part of the workflow.

--gap_accessions /Users/pleyte/Documents/office/ohsu/cgd/taskNotes/bio818-RefSeqMIsalignment/Final_Review/1_HowManyReportedTranscriptsWithGaps/gap_accessions_from_gff_and_uta.csv
--cgd_accessions /Users/pleyte/Documents/office/ohsu/cgd/taskNotes/bio818-RefSeqMIsalignment/Final_Review/1_HowManyReportedTranscriptsWithGaps/cgd_distinct_transcript_accessions.csv
--cgd_variant_transcripts /Users/pleyte/Documents/office/ohsu/cgd/taskNotes/transcript_effects_testing/cgd_prod_2020-2026_reported_export/cgd_prod_2020-2026_reported_export.csv
--tfx_nomenclature /Users/pleyte/git/refseq-indel-nomenclature-comparison/nextflow/results/tfx_nomenclature.csv
--accession_index /tmp/gff/GCF_000001405.25_GRCh37.p13_genomic_accession_index.parquet
--out_xlsx /Users/pleyte/Documents/office/ohsu/cgd/taskNotes/bio818-RefSeqMIsalignment/Final_Review/1_HowManyReportedTranscriptsWithGaps/cgd_transcripts_with_gaps.xlsx 

@author: pleyte
'''
import argparse
import pandas as pd
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def compare_transcripts(gap_file, cgd_transcript_file, cgd_variants_file, tfx_nomenclature_file, accession_index_file, output_xlsx):
    # 1. Load the data
    gap_df = pd.read_csv(gap_file)
    cgd_tx_df = pd.read_csv(cgd_transcript_file)
    variants_df = pd.read_csv(cgd_variants_file)
    tfx_df = pd.read_csv(tfx_nomenclature_file)
    accession_index_df = pd.read_parquet(accession_index_file).reset_index()
    
    for df in [variants_df, tfx_df]:
        df['chromosome'] = df['chromosome'].astype(str)
        
    # 3. Build Transcript Comparison (Sheet 1)
    tx_comparison = _get_tx_comparison_df(gap_df, cgd_tx_df)
    
    # 4. Filter for Gapped Transcripts
    # We want any transcript where EITHER match is True to trigger the variant pull
    gapped_tx_list = tx_comparison[
        (tx_comparison['Matches_Gap_Base_Accession']) | 
        (tx_comparison['Matches_Gap_Exact_Version'])
    ]['CGD_Transcript'].unique()

    # 5. Filter the Variants Dataframe (Sheet 2)
    # Using 'cdna_transcript' as the join key from the clinical file
    affected_variants_df = variants_df[variants_df['cdna_transcript'].isin(gapped_tx_list)].copy()
    affected_variants_df['chromosome'] = affected_variants_df['chromosome'].astype(str).str.replace('chr', '', case=False)
  
    tfx_comparison_df = _get_tfx_comparison_df(tfx_df, affected_variants_df, accession_index_df)
    
    # Summary Stats
    print(f"--- Variant Analysis Complete ---")
    print(f"Total Unique Transcripts checked: {len(tx_comparison)}")
    print(f"Transcripts found with Gaps:      {len(gapped_tx_list)}")
    print(f"Total Variant Rows Filtered:      {len(affected_variants_df)}")
    print(f"Tfx Mismatches Detected:          {tfx_comparison_df['nomenclature_mismatch'].sum()}")
    
    _write(output_xlsx, tx_comparison, affected_variants_df, tfx_comparison_df)

def _get_tx_comparison_df(gap_df, cgd_tx_df):
    """
    Dataframe that indicates whether an accession is known to have gaps
    """
    # Identify accession columns
    gap_col = 'accession' 
    assert 'accession' in gap_df.columns 
    cgd_tx_col = 'accession' 
    assert 'accession' in cgd_tx_df.columns 

    # 2. Prepare lookup sets for speed
    # Exact: 'NM_000123.4'
    gap_set_exact = set(gap_df[gap_col].unique())
    # Versionless: 'NM_000123'
    gap_set_versionless = set(gap_df[gap_col].str.split('.').str[0].unique())

    tx_comparison = pd.DataFrame()
    tx_comparison['CGD_Transcript'] = cgd_tx_df[cgd_tx_col]
    
    # Create the check flags
    cgd_tx_stripped = tx_comparison['CGD_Transcript'].str.split('.').str[0]
    tx_comparison['Matches_Gap_Base_Accession'] = cgd_tx_stripped.isin(gap_set_versionless)
    tx_comparison['Matches_Gap_Exact_Version'] = tx_comparison['CGD_Transcript'].isin(gap_set_exact)
    return tx_comparison
    
def _get_tfx_comparison_df(tfx_df, affected_variants_df, accession_index_df):
    """
    Join CGD and Tfx rows. Highlighting  
    """
    tfx_df['chromosome'] = tfx_df['chromosome'].astype(str)
    
    # Initial Merge with Tfx
    tfx_comparison_df = pd.merge(
        affected_variants_df,
        tfx_df,
        left_on=['chromosome', 'position_start', 'reference_base', 'variant_base', 'cdna_transcript'],
        right_on=['chromosome', 'position', 'reference', 'alt', 'cdna_transcript'],
        how='left'
    )
    
    # Join with Accession Index Parquet data
    tfx_comparison_df = pd.merge(
        tfx_comparison_df,
        accession_index_df[['accession', 'gap', 'target_start', 'target_end']],
        left_on='cdna_transcript',
        right_on='accession',
        how='left'
    )
    
    # Logic for nomenclature_mismatch: 
    # Use pd.NA friendly comparison or fillna to handle missing data
    c_mismatch = tfx_comparison_df['genotype_cdna'].fillna('') != tfx_comparison_df['c_dot.tfx'].fillna('')
    p_mismatch = tfx_comparison_df['genotype_amino_acid_onel'].fillna('') != tfx_comparison_df['p_dot1.tfx'].fillna('')    
        
    tfx_comparison_df['nomenclature_mismatch'] = (c_mismatch | p_mismatch).astype(int)    
    
    # Reorder columns to make it easier to compare them 
    leader_cols = [
        'genomic_variant', 'chromosome', 'position_start', 'reference_base', 'variant_base',
        'exon', 'exon.tfx', 'cdna_transcript', 'protein_transcript', 'protein_transcript.tfx',
        'genotype_cdna', 'c_dot.tfx', 'genotype_amino_acid_onel', 'p_dot1.tfx', 'p_dot3.tfx',
        'nomenclature_mismatch', 'gap', 'target_start', 'target_end'
    ]
    remaining_cols = [c for c in tfx_comparison_df.columns if c not in leader_cols]
    final_col_order = [c for c in leader_cols if c in tfx_comparison_df.columns] + remaining_cols
    tfx_comparison_df = tfx_comparison_df[final_col_order]
    return tfx_comparison_df


def _write(output_xlsx, tx_comparison, affected_variants_df, tfx_comparison_df):
    """
    """
    # 6. Save to Excel
    with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
        # Sheet 1: list of transcripts and whether they have gap issues
        tx_comparison.to_excel(writer, sheet_name='Transcript_Check', index=False)
        
        # Sheet 2: All of CGD's rows for variants on those gapped transcripts
        affected_variants_df.to_excel(writer, sheet_name='Affected_Variants', index=False)
        
        # Sheet 3: Join tfx nomenclature
        tfx_comparison_df.to_excel(writer, sheet_name='CGD_Tfx_Comparison', index=False)
        workbook = writer.book
        apply_comparison_format(workbook['CGD_Tfx_Comparison'], tfx_comparison_df, 'genotype_cdna', 'c_dot.tfx')
        apply_comparison_format(workbook['CGD_Tfx_Comparison'], tfx_comparison_df, 'genotype_amino_acid_onel', 'p_dot1.tfx')
    
        print(f"Results written to: {output_xlsx}")
        

    
def apply_comparison_format(worksheet, tfx_comparison_df, col_a_name, col_b_name):
    last_row = len(tfx_comparison_df) + 1
    
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid') # Light Green
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')   # Light Red
    
    try:
        # Get Excel column letters
        idx_a = tfx_comparison_df.columns.get_loc(col_a_name) + 1
        idx_b = tfx_comparison_df.columns.get_loc(col_b_name) + 1
        let_a = get_column_letter(idx_a)
        let_b = get_column_letter(idx_b)

        # Range for both columns (e.g., J2:K100)
        fmt_range = f'{let_a}2:{let_b}{last_row}'

        # Rule 1: Equal (Green). Formula uses $ to lock columns but relative rows
        worksheet.conditional_formatting.add(fmt_range,
            FormulaRule(formula=[f'${let_a}2=${let_b}2'], stopIfTrue=True, fill=green_fill))
        
        # Rule 2: Not Equal (Red)
        worksheet.conditional_formatting.add(fmt_range,
            FormulaRule(formula=[f'${let_a}2<>${let_b}2'], stopIfTrue=True, fill=red_fill))
    except KeyError:
        pass

def main():
    
    parser = argparse.ArgumentParser(description='Identify CGD''s reported transcripts that are known to have reference gaps')
    parser.add_argument('--gap_accessions', help="Accessions known to have reference mismatch (csv)", required=True)
    parser.add_argument('--cgd_accessions', help="Accessions associated with reported variants in CGD (csv)", required=True)
    parser.add_argument('--cgd_variant_transcripts', help="Variants and transcript accessions exported from CGD (csv)", required=True)
    parser.add_argument('--tfx_nomenclature', help="Tfx variant transcripts and nomenclature (csv)", required=True)
    parser.add_argument('--accession_index', help="Gff accession index with gap info (parquet)", required=True)
    parser.add_argument('--out_xlsx', help="Save results to file (xlsx)", required=True)
    
    args = parser.parse_args()
    
    compare_transcripts(args.gap_accessions, 
                        args.cgd_accessions,
                        args.cgd_variant_transcripts, 
                        args.tfx_nomenclature,
                        args.accession_index,
                        args.out_xlsx)
        
if __name__ == '__main__':
    main()